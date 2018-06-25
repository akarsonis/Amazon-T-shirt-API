from bs4 import BeautifulSoup
import requests
from fake_useragent import UserAgent

import re

import openpyxl
from openpyxl import Workbook

#Initial Variables

wb = Workbook()
ws = wb.active

ua = UserAgent()
print(ua.chrome)
header = {'User-Agent':str(ua.Chrome)}

#Main loop

def data_from_amazon(max_page, keyword):
    
    if ' ' in keyword:
        keyword = keyword.replace(' ', '+')

    single_company_count = 1
    single_product_count = 1
    fractional_price_count = 1
    full_price_count = 1
    main_link_count = 2
    
    page = 1
    
    while page < max_page + 1:
        url = 'https://www.amazon.com/s/ref=sr_pg_2?rh=n%3A7141123011%2Cn%3A7147445011%2Ck%3Aanimal&page=' + str(page) +'&hidden-keywords=ORCA&keywords=' + str(keyword) +'&ie=UTF8&qid=1529754410'
        htmlContent = requests.get(url, headers=header)
        soup = BeautifulSoup(htmlContent.content, 'html.parser')
        htmlContent = soup.prettify()
        
        #Price

        span_price = soup.find_all('span', {'class' : 'sx-price-whole'})
        for single_full_price in span_price:
            single_full_price = single_full_price.string.replace("<", "")
            ws.cell(row=full_price_count, column=3).value = single_full_price
            full_price_count += 1

        sup_price = soup.find_all('sup', {'class' : 'sx-price-fractional'})
        for single_fractional_price in sup_price:
            single_fractional_price = single_fractional_price.string.replace("<", "")
            ws.cell(row=fractional_price_count, column=3).value = ws.cell(row=fractional_price_count, column=3).value + '.' + single_fractional_price
            fractional_price_count += 1
        
        #Company name
        
        spans_company_names = soup.find_all('span', {'class' : 'a-color-secondary s-overflow-ellipsis s-size-mild'})
        for single_company_name in spans_company_names:
            single_company_name = single_company_name.string.replace("<", "")
            ws.cell(row=single_company_count, column=2).value = single_company_name
            single_company_count += 1
        
        #Product name
        
        a_single_product = soup.find_all('a', {'class' : 'a-link-normal s-access-detail-page s-overflow-ellipsis s-color-twister-title-link a-text-normal'})
        for single_product_name in a_single_product:
            single_product_name = single_product_name.string.replace("<", "")
            ws.cell(row=single_product_count, column=1).value = single_product_name
            single_product_count += 1
            
        #Links
        
        main_link_count -= 1
        
        a_link = soup.find_all('a', {'class' : 'a-link-normal a-text-normal'})
        for single_main_link in a_link:
            single_main_link = str(single_main_link)

            if 'www.amazon.com' in single_main_link and '\"><img alt=\"' in single_main_link:
                single_main_link = single_main_link.replace('<a class="a-link-normal a-text-normal" href="', '')
                single_main_link = single_main_link.split('\"><img alt=\"', 1)[0]
                
                #Now accessing each link separately
                
                product_html_Content = requests.get(single_main_link, headers=header)
                product_soup = BeautifulSoup(product_html_Content.content, 'html.parser')
                product_html_Content = product_soup.prettify()                
                
                #Retrieving descriptions

                p_description = product_soup.find_all(['p'])
                p_description.append('No description')
                
                for single_description in p_description:
                    single_description = str(single_description)
                    if 'class=' not in single_description and 'Sponsored Products are advertisements' not in single_description:
                        single_description = single_description.replace("<p>", "")
                        single_description = single_description.replace("</p>", "")
                        single_description = re.sub(r'\n\s*\n', r'\n\n', single_description.strip(), flags=re.M)
                        del p_description[-1]
                        print(single_description)
                        print(main_link_count)
                        ws.cell(row=main_link_count, column=4).value = single_description
                        main_link_count -= 3

            main_link_count += 1
            
                
            
        page += 1
    
    wb.save('Amazon API.xlsx')

data_from_amazon(2, 'animal')
